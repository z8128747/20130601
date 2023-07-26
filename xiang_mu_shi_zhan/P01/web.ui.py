import openpyxl

from xiang_mu_shi_zhan.P01.key import open_brower, Keys


def arguments(values):
    data = dict()
    if values:
        # 基于;号拆分
        str_tem = values.split(';')

        for tem in str_tem:
            # print(tem)
            # 基于=号切分，切1次，因为xpath定位也会出现=号
            t = tem.split('=', 1)
            # print(t)
            data[t[0]] = t[1]

    else:
        pass
    return data


excel = openpyxl.load_workbook('..//date/自动化测试用例2.xlsx')
sheet1 = excel['Sheet1']
for values in sheet1.values:
    if type(values[0]) is int:
        # 用例描述可以用作日志
        print(f'----------正在执行{values[3]}----------')
        # print(values)
        data = arguments(values[2])

        if values[1] == 'open_brower':
            key = Keys(**data)
        elif 'assert' in values[1]:

            status=getattr(key, values[1])(expected=values[4], **data)
            print(status)
            if status:
                sheet1.cell(row=values[0] + 1, column=6).value = 'Pass'
            else:
                sheet1.cell(row=values[0] + 1, column=6).value = 'False'
            #     保存Excel,放在这里确保每次运行都保存，防止报错没保存
            excel.save('../date/自动化测试用例2.xlsx')



        else:
            getattr(key, values[1])(**data)

